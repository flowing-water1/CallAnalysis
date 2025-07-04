import streamlit as st
import os
import asyncio
import logging
import concurrent.futures
from io import BytesIO
import re
import openpyxl
from datetime import date, datetime
import pytz
from config import LOGGING_CONFIG, EXCEL_CONFIG, DATABASE_CONFIG, get_current_db_config
from database_utils import SyncDatabaseManager
from Audio_Recognition import (
    process_file,
    process_all_files
)
from Identify_Roles import (
    identify_roles,
    format_conversation_with_roles
)
from Analyze_Conversation import analyze_conversation_with_roles
from Analyze_Summary import analyze_summary
from LLM_Workflow import llm_workflow
from extract_utils import extract_all_conversation_data, extract_all_summary_data, parse_filename_intelligently
import json

# 配置日志输出
logging.basicConfig(
    level=getattr(logging, LOGGING_CONFIG["level"]), 
    format=LOGGING_CONFIG["format"]
)
logger = logging.getLogger(__name__)

def run_async_process(coro):
    """专门用于运行process_all_files的异步包装器"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()

@st.dialog(title="欢迎使用通话分析工具！", width="large")
def tutorial():
    st.markdown("## 📚 使用教程")
    st.markdown("### ⚠️ 重要格式要求")
    st.markdown(
        "上传文件的格式必须是 :red[**\"公司名称-联系人-电话号码\"**] 的形式。中间有无空格不影响，但必须使用 :red[**\"-\"**] 作为分隔符。（此格式要求将在后续版本中优化）",
        unsafe_allow_html=True)

    st.markdown("### 使用流程")
    st.markdown("#### 0️⃣ 选择名字")
    st.markdown("选择自己的名字")
    st.image("tutorial/名字.png")

    st.markdown("#### 1️⃣ 上传文件")
    st.markdown("点击下方按钮上传您的通话录音文件：")
    st.image("tutorial/上传文件按钮.png")
    st.markdown(":green[✅] 支持批量上传多个文件")
    st.image("tutorial/上传文件.png")

    st.markdown("#### 2️⃣ 确认上传状态")
    st.markdown("成功上传后，您将看到如下界面：")
    st.image("tutorial/上传之后的样子.png")

    st.markdown("#### 2️⃣-1️⃣ 倘若今天已经上传过一次")
    st.markdown("可以选择「覆盖数据库中的信息」，也可以将当前的数据「追加到数据库中」")
    st.image("tutorial/已有记录.png")
    st.markdown("#### 3️⃣ 开始分析流程")
    st.markdown("点击 :blue[**\"开始分析\"**] 按钮启动处理：")
    st.image("tutorial/开始分析.png")

    st.markdown("#### 4️⃣ 等待处理完成")
    st.markdown("系统正在处理中，请保持页面打开。您可以暂时切换到其他工作，处理完成后回来查看结果。")

    st.markdown("#### 5️⃣ 查看分析结果")
    st.image("tutorial/最终结果.png")

    st.markdown("#### 6️⃣ 导出分析报告")
    st.markdown("您可以下载：")
    st.markdown("• :blue[完整分析报告] - 包含所有通话记录和详细分析")
    st.markdown("• :green[电话开拓分析表] - 自动填写好的分析数据表格")

    st.markdown("表格中已自动填写好对应数据项：")
    st.image("tutorial/分析结果表格.png")

    st.markdown("分析报告采用Markdown格式，建议使用Markdown编辑器打开以获得最佳阅读体验：")
    st.image("tutorial/分析结果文档.png")

    st.markdown("### ❓ 如何关闭本教程")
    st.markdown("点击对话框外任意位置，或滚动至顶部点击右上角的'❌'即可关闭本教程。")


# Streamlit界面
st.set_page_config(page_title="分析通话记录Demo", page_icon="📞")

# 使用列布局让标题和按钮在同一行
title_col, button_col = st.columns([5, 1.2])

# 在第一列放置标题
with title_col:
    st.title("分析通话记录📞")

# 在第二列放置帮助按钮
with button_col:
    if st.button("📚 查看教程", help="点击查看详细使用教程"):
        # 当按钮被点击时，触发教程对话框
        tutorial()

# 初始化session state
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = None
if 'combined_report' not in st.session_state:
    st.session_state.combined_report = None
if 'summary_analysis' not in st.session_state:
    st.session_state.summary_analysis = None
if 'analysis_completed' not in st.session_state:
    st.session_state.analysis_completed = False  # 用来标记分析是否完成
if 'tutorial_shown' not in st.session_state:
    st.session_state.tutorial_shown = False
if 'db_manager' not in st.session_state:
    st.session_state.db_manager = None
if 'salesperson_id' not in st.session_state:
    st.session_state.salesperson_id = None
if 'salesperson_name' not in st.session_state:
    st.session_state.salesperson_name = None
if 'upload_choice' not in st.session_state:
    st.session_state.upload_choice = None

# 初始化数据库管理器
def get_db_manager():
    """获取数据库管理器实例（不使用缓存以避免连接问题）"""
    # 移除缓存装饰器，每次都创建新实例，避免连接断开问题
    return SyncDatabaseManager(get_current_db_config())

# 仅在第一次加载页面且教程未显示过时显示教程
if not st.session_state.tutorial_shown:
    tutorial()
    st.session_state.tutorial_shown = True

# 销售人员选择区域
st.markdown("### 🙋🏻‍♂️ 请选择您的姓名")

# 获取销售人员列表
try:
    db_manager = get_db_manager()
    salespersons = db_manager.get_salespersons()
    salesperson_names = ["请选择..."] + [sp['name'] for sp in salespersons]
    
    # 销售人员下拉选择框
    selected_name = st.selectbox(
        "选择销售人员",
        options=salesperson_names,
        key="salesperson_select",
        help="请从下拉列表中选择您的姓名"
    )
    
    # 如果选择了有效的销售人员
    if selected_name != "请选择...":
        # 查找对应的销售人员ID
        selected_person = next((sp for sp in salespersons if sp['name'] == selected_name), None)
        if selected_person:
            st.session_state.salesperson_id = selected_person['id']
            st.session_state.salesperson_name = selected_person['name']
            st.success(f"已选择：{selected_name}")
        else:
            st.error("选择的销售人员不存在")
    else:
        st.session_state.salesperson_id = None
        st.session_state.salesperson_name = None
        
except Exception as e:
    st.error(f"获取销售人员列表失败：{str(e)}")
    st.info("请检查数据库连接是否正常")

# 只有选择了销售人员才能选择处理模式
if st.session_state.salesperson_id:
    st.markdown("---")
    st.markdown("### 📋 选择处理模式")
    
    # 初始化处理模式状态
    if 'processing_mode' not in st.session_state:
        st.session_state.processing_mode = "📞 录音文件"
    
    # 模式选择
    processing_mode = st.radio(
        "请选择要处理的内容类型",
        options=["📞 录音文件", "📸 聊天截图"],
        horizontal=True,
        help="录音文件：分析音频通话记录；聊天截图：识别微信聊天中的通话时长信息"
    )
    
    # 更新session state
    st.session_state.processing_mode = processing_mode
    
    st.markdown("---")
    
    if processing_mode == "📞 录音文件":
        # 现有的录音文件上传流程
        st.markdown("### 📁 上传通话录音文件")
        
        uploaded_files = st.file_uploader(
            "请上传通话录音文件",
            type=['wav', 'mp3', 'm4a', 'ogg', 'aac'],
            accept_multiple_files=True,
            help="支持 WAV、MP3、M4A、OGG、AAC 格式的音频文件"
        )
        uploaded_images = None  # 确保图片变量为空
        
    else:
        # 新的图片上传流程
        st.markdown("### 📸 上传微信通话截图")
        
        uploaded_images = st.file_uploader(
            "请上传微信聊天截图",
            type=['jpg', 'jpeg', 'png', 'bmp'],
            accept_multiple_files=True,
            help="请上传包含通话时长信息的微信聊天截图"
        )
        uploaded_files = None  # 确保音频变量为空

        # 图片预览
        if uploaded_images:
            from image_utils import create_image_preview_grid
            create_image_preview_grid(uploaded_images, columns=3)
else:
    st.warning("⚠️ 请先选择您的姓名后才能上传文件")
    uploaded_files = None
    uploaded_images = None

if uploaded_files and not st.session_state.analysis_completed:
    st.write("已上传的文件:")
    for file in uploaded_files:
        st.write(f"- {file.name}")
    
    # 🔍 重复文件检测
    try:
        # 提取文件名列表
        filenames = [file.name for file in uploaded_files]
        
        # 初始化数据库管理器
        db_manager = get_db_manager()
        
        # 检测重复文件
        duplicate_check = db_manager.check_duplicate_filenames(
            st.session_state.salesperson_id, 
            filenames,
            days_back=30  # 检测最近30天
        )
        
        # 显示检测结果
        if duplicate_check["duplicates"] or duplicate_check["new_files"]:
            st.markdown("---")
            st.markdown("### 🔍 文件重复检测结果")
            
            # 显示新文件
            if duplicate_check["new_files"]:
                st.success(f"✅ **新文件 ({len(duplicate_check['new_files'])} 个)**：将正常处理")
                with st.expander("📋 查看新文件列表", expanded=False):
                    for new_file in duplicate_check["new_files"]:
                        st.write(f"- {new_file}")
            
            # 显示重复文件
            if duplicate_check["duplicates"]:
                st.warning(f"⚠️ **重复文件 ({len(duplicate_check['duplicates'])} 个)**：已自动跳过")
                with st.expander("📋 查看重复文件详情", expanded=True):
                    for dup in duplicate_check["duplicates"]:
                        days_text = "今天" if dup["days_ago"] == 0 else f"{dup['days_ago']} 天前"
                        st.write(f"- **{dup['filename']}**")
                        st.write(f"  └─ 完全相同的文件已于 {days_text} ({dup['last_upload_date']}) 上传过")
                
                # 如果所有文件都是重复的，提前结束
                if not duplicate_check["new_files"]:
                    st.info("💡 所有文件都是重复文件，无需处理。请选择其他文件后重新上传。")
                    st.stop()  # 停止执行后续代码
        
        # 过滤掉重复文件，只处理新文件
        if duplicate_check["new_files"]:
            # 重新构建 uploaded_files 列表，只包含新文件
            new_uploaded_files = [
                file for file in uploaded_files 
                if file.name in duplicate_check["new_files"]
            ]
            
            # 如果有文件被过滤掉，显示过滤后的文件数量
            if len(new_uploaded_files) < len(uploaded_files):
                st.info(f"📝 已过滤重复文件，将处理 {len(new_uploaded_files)} 个新文件")
            
            # 使用过滤后的文件列表继续后续处理
            uploaded_files = new_uploaded_files
        else:
            # 如果没有新文件，停止处理
            st.stop()
            
    except Exception as e:
        st.error(f"检测重复文件时出错：{str(e)}")
        st.info("将跳过重复检测，继续处理所有文件")
    
    # 检查是否已有今日记录
    today = date.today()
    
    try:
        has_existing_record = db_manager.check_daily_record_exists(
            st.session_state.salesperson_id, 
            today
        )
        
        if has_existing_record and st.session_state.upload_choice is None:
            st.warning(f"⚠️ {st.session_state.salesperson_name} 今天已有上传记录")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("覆盖现有数据", type="primary"):
                    st.session_state.upload_choice = "overwrite"
                    st.rerun()
            with col2:
                if st.button("追加到现有数据"):
                    st.session_state.upload_choice = "append"
                    st.rerun()
            with col3:
                if st.button("取消本次上传", type="secondary"):
                    st.session_state.upload_choice = "cancel"
                    st.session_state.analysis_completed = True
                    st.rerun()
                    
        # 如果选择了取消，不显示分析按钮
        if st.session_state.upload_choice == "cancel":
            st.info("已取消本次上传")
        elif not has_existing_record or st.session_state.upload_choice in ["overwrite", "append"]:
            if st.button("开始分析", key="start_analysis"):
                with st.spinner("正在处理文件..."):
                    progress_placeholder = st.empty()
                    # 保存上传的文件到临时文件夹
                    temp_files = []
                    for uploaded_file in uploaded_files:
                        # 确保临时文件夹存在
                        os.makedirs("temp", exist_ok=True)
                        temp_path = os.path.join("temp", f"temp_{uploaded_file.name}")
                        with open(temp_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        temp_files.append(temp_path)

                    try:
                        results = run_async_process(process_all_files(temp_files, progress_placeholder))
                        st.session_state.analysis_results = results

                        # 生成汇总分析并保存，同时更新进度条（汇总分析占 20%）
                        phase_text = progress_placeholder.empty()
                        phase_text.markdown("**🔄 正在生成汇总分析...**")
                        progress_bar = progress_placeholder.progress(0.9)
                        st.session_state.summary_analysis = analyze_summary([res for res in results if res["status"] == "success"])
                        progress_bar.progress(1.0)
                        phase_text.markdown("**✅ 所有文件处理完成！**")

                        # 生成完整报告并保存
                        combined_report = ""
                        for idx, res in enumerate(results, 1):
                            if res["status"] == "success" and res["analysis_result"].get("status") == "success":
                                combined_report += f"\n\n{'=' * 50}\n对话记录 {idx}：\n{'=' * 50}\n\n"
                                combined_report += res["analysis_result"]["formatted_text"]
                                combined_report += f"\n\n{'=' * 50}\n分析结果 {idx}：\n{'=' * 50}\n\n"
                                combined_report += res["analysis_result"]["analysis"]

                        combined_report += f"\n\n{'=' * 50}\n汇总分析报告：\n{'=' * 50}\n\n"
                        combined_report += st.session_state.summary_analysis
                        st.session_state.combined_report = combined_report
                        
                        # 保存分析结果到数据库
                        phase_text.markdown("**💾 正在保存分析结果到数据库...**")
                        try:
                            # 准备数据
                            call_details_list = []
                            for res in results:
                                if res["status"] == "success" and res["analysis_result"].get("status") == "success":
                                    # 解析文件名
                                    file_name = os.path.basename(res["file_path"])
                                    file_name = re.sub(r'^temp_', '', file_name)
                                    file_name_without_ext = os.path.splitext(file_name)[0]
                                    
                                    # 使用智能文件名解析
                                    company_name, contact_person, phone_number = parse_filename_intelligently(file_name_without_ext)
                                    
                                    # 获取对话文本
                                    conversation_text = res["analysis_result"]["formatted_text"]
                                    
                                    # 提取分析数据
                                    analysis_text = res["analysis_result"]["analysis"]
                                    extracted_data = extract_all_conversation_data(analysis_text)
                                    
                                    # 正确获取评分（用于统计，不用于有效性判断）
                                    score = None
                                    if extracted_data["score"]:
                                        try:
                                            score = float(extracted_data["score"])
                                        except ValueError:
                                            pass
                                    
                                    # 🚀 修复：使用正确的时间判断逻辑（>= 60秒）
                                    # 从 Audio_Recognition.py 的结果中获取已经计算好的 is_valid_call
                                    is_effective = res.get("is_valid_call", False)
                                    
                                    # 准备分析结果的JSON格式
                                    analysis_result_json = {
                                        "roles": res["analysis_result"].get("roles", {}),
                                        "analysis": analysis_text,
                                        "extracted_data": extracted_data,
                                        "suggestions": extracted_data["suggestion"]
                                    }
                                    
                                    # 准备单条通话详情（使用新字段名）
                                    call_detail = {
                                        'original_filename': file_name,
                                        'company_name': company_name,
                                        'contact_person': contact_person,
                                        'phone_number': phone_number,
                                        'conversation_text': conversation_text,
                                        'analysis_text': analysis_text,
                                        'score': score,
                                        'is_effective': is_effective,
                                        'suggestions': extracted_data.get("suggestion", "")
                                    }
                                    
                                    call_details_list.append(call_detail)
                            
                            # 保存到数据库
                            save_success = db_manager.save_analysis_data(
                                st.session_state.salesperson_id,
                                call_details_list,
                                st.session_state.summary_analysis,
                                st.session_state.upload_choice
                            )
                            
                            if save_success:
                                phase_text.markdown("**✅ 分析结果已成功保存到数据库！**")
                            else:
                                st.warning("分析结果保存到数据库时出现问题，但您仍可以下载报告")
                        except Exception as db_error:
                            st.error(f"保存到数据库失败：{str(db_error)}")
                            st.info("您仍然可以下载分析报告")

                        st.session_state.analysis_completed = True  # 标记分析完成

                    except Exception as e:
                        st.error(f"处理过程中出现错误：{str(e)}")
                    finally:
                        # 清理临时文件
                        for temp_file in temp_files:
                            if os.path.exists(temp_file):
                                os.remove(temp_file)
                        # 清理转换生成的临时文件（保留有价值的转换文件信息）
                        converted_files_to_cleanup = []
                        for res in results:
                            if res["status"] == "success" and "conversion_info" in res:
                                conversion_info = res["conversion_info"]
                                if conversion_info.get("conversion_success", False):
                                    converted_file_path = conversion_info.get("converted_file_path")
                                    if converted_file_path and os.path.exists(converted_file_path):
                                        converted_files_to_cleanup.append(converted_file_path)
                        
                        # 延迟清理转换文件，给用户查看的时间
                        if converted_files_to_cleanup:
                            logging.info(f"发现 {len(converted_files_to_cleanup)} 个转换文件，将在session结束时清理")
                            # 将转换文件列表保存到session state中，用于后续清理
                            if 'converted_files_cleanup' not in st.session_state:
                                st.session_state.converted_files_cleanup = []
                            st.session_state.converted_files_cleanup.extend(converted_files_to_cleanup)
                        
                        # 尝试删除临时文件夹（只删除空文件夹）
                        try:
                            os.rmdir("temp")
                        except OSError:
                            pass  # 如果文件夹不为空或不存在，忽略错误

    except Exception as e:
        st.error(f"检查数据库记录时出错：{str(e)}")
        has_existing_record = False
        # 即使出错也允许继续分析
        if st.button("开始分析", key="start_analysis"):
            with st.spinner("正在处理文件..."):
                progress_placeholder = st.empty()

# 处理图片文件的逻辑（新增）
if uploaded_images and not st.session_state.analysis_completed:
    st.write("已上传的图片:")
    for img in uploaded_images:
        st.write(f"- {img.name}")
    
    # 检查是否已有今日记录
    today = date.today()
    
    try:
        db_manager = get_db_manager()
        has_existing_record = db_manager.check_daily_record_exists(
            st.session_state.salesperson_id, 
            today
        )
        
        if has_existing_record and st.session_state.upload_choice is None:
            st.warning(f"⚠️ {st.session_state.salesperson_name} 今天已有上传记录")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("覆盖现有数据", type="primary", key="img_overwrite"):
                    st.session_state.upload_choice = "overwrite"
                    st.rerun()
            with col2:
                if st.button("追加到现有数据", key="img_append"):
                    st.session_state.upload_choice = "append"
                    st.rerun()
            with col3:
                if st.button("取消本次上传", type="secondary", key="img_cancel"):
                    st.session_state.upload_choice = "cancel"
                    st.session_state.analysis_completed = True
                    st.rerun()
                    
        # 如果选择了取消，不显示分析按钮
        if st.session_state.upload_choice == "cancel":
            st.info("已取消本次上传")
        elif not has_existing_record or st.session_state.upload_choice in ["overwrite", "append"]:
            
            # 🔍 第一层去重检查：文件名重复检测
            st.markdown("### 🔍 检查文件重复情况")
            
            # 执行去重检查（只在第一次执行）
            if 'image_duplicate_result' not in st.session_state:
                with st.spinner("正在检查文件名重复情况..."):
                    try:
                        from Image_Recognition import check_image_duplicates
                        
                        duplicate_result = check_image_duplicates(
                            uploaded_images, 
                            st.session_state.salesperson_id, 
                            db_manager
                        )
                        
                        st.session_state.image_duplicate_result = duplicate_result
                        
                    except Exception as e:
                        st.error(f"去重检查时出现错误：{str(e)}")
                        # 出错时默认所有文件都是新文件
                        st.session_state.image_duplicate_result = {
                            "has_duplicates": False,
                            "duplicates": [],
                            "new_files": [img.name for img in uploaded_images],
                            "duplicate_files": [],
                            "clean_files": uploaded_images,
                            "total_images": len(uploaded_images),
                            "duplicate_count": 0,
                            "new_count": len(uploaded_images),
                            "error": str(e)
                        }
            
            # 显示去重分析结果和用户选择界面
            duplicate_result = st.session_state.image_duplicate_result
            
            if 'image_user_choice' not in st.session_state:
                from image_utils import display_duplicate_analysis
                
                user_choice = display_duplicate_analysis(duplicate_result)
                
                if user_choice:
                    st.session_state.image_user_choice = user_choice
                    st.rerun()
            
            else:
                # 用户已做出选择，显示选择结果
                user_choice = st.session_state.image_user_choice
                
                if user_choice == "cancel":
                    st.error("❌ 已取消图片上传")
                    st.session_state.analysis_completed = True
                
                elif user_choice in ["proceed", "skip_duplicates", "force_all"]:
                    # 根据用户选择过滤图片
                    from Image_Recognition import filter_duplicate_images
                    
                    if user_choice == "proceed":
                        # 没有重复文件，直接处理所有图片
                        filtered_images = uploaded_images
                        st.success(f"✅ 准备处理 {len(filtered_images)} 张新图片")
                    else:
                        # 有重复文件，根据用户选择过滤
                        filtered_images = filter_duplicate_images(
                            uploaded_images, 
                            duplicate_result, 
                            user_choice
                        )
                        
                        if user_choice == "skip_duplicates":
                            st.info(f"📝 将处理 {len(filtered_images)} 张新图片，跳过 {duplicate_result.get('duplicate_count', 0)} 张重复图片")
                        elif user_choice == "force_all":
                            st.warning(f"⚠️ 将强制处理所有 {len(filtered_images)} 张图片（包括重复文件）")
                    
                    # 保存过滤后的图片列表
                    st.session_state.filtered_images = filtered_images
                    
                    # 显示开始识别按钮
                    if len(filtered_images) > 0:
                        if st.button("开始识别", key="start_image_analysis", type="primary"):
                            with st.spinner("正在识别图片中的通话信息..."):
                                progress_placeholder = st.empty()
                                
                                try:
                                    # 导入图片识别模块
                                    from Image_Recognition import process_image_batch, prepare_database_update_data
                                    from image_utils import display_processing_summary, handle_image_processing_errors
                                    
                                    # 异步处理图片
                                    def update_progress(progress, message):
                                        progress_placeholder.progress(progress)
                                        progress_placeholder.markdown(f"**{message}**")
                                    
                                    # 处理图片批次（使用过滤后的图片列表）
                                    def run_image_process():
                                        loop = asyncio.new_event_loop()
                                        asyncio.set_event_loop(loop)
                                        try:
                                            return loop.run_until_complete(
                                                process_image_batch(filtered_images, update_progress)
                                            )
                                        finally:
                                            loop.close()
                                    
                                    processing_results = run_image_process()
                                    
                                    # 🤖 第二层去重检查：智能内容去重
                                    if processing_results.get('all_calls') and len(processing_results['all_calls']) > 0:
                                        st.markdown("### 🤖 智能内容去重检查")
                                        
                                        with st.spinner("正在进行智能去重分析..."):
                                            # 获取现有数据库记录进行比较
                                            existing_calls = db_manager.get_recent_call_records(
                                                st.session_state.salesperson_id, 
                                                days_back=30
                                            )
                                            
                                            # 智能去重检测
                                            from Image_Recognition import smart_duplicate_detection
                                            detection_result = smart_duplicate_detection(
                                                processing_results['all_calls'], 
                                                existing_calls
                                            )
                                            
                                            # 显示去重结果
                                            from image_utils import display_smart_duplicate_result
                                            should_continue = display_smart_duplicate_result(detection_result)
                                            
                                            if should_continue:
                                                # 更新处理结果，只保留非重复的记录
                                                processing_results['all_calls'] = detection_result['processed_calls']
                                                
                                                # 重新计算统计数据
                                                total_calls = len(processing_results['all_calls'])
                                                effective_calls = sum(1 for call in processing_results['all_calls'] 
                                                                    if call.get('is_effective', False))
                                                
                                                processing_results['total_calls_found'] = total_calls
                                                processing_results['effective_calls_found'] = effective_calls
                                            else:
                                                # 没有新记录需要处理
                                                processing_results['all_calls'] = []
                                                processing_results['total_calls_found'] = 0
                                                processing_results['effective_calls_found'] = 0
                                            
                                            st.markdown("---")
                                    
                                    # 显示处理结果摘要
                                    display_processing_summary({
                                        'total': processing_results['total_images'],
                                        'success': processing_results['successful_images'], 
                                        'failed': processing_results['failed_images'],
                                        'calls_found': processing_results['total_calls_found'],
                                        'effective_calls': processing_results['effective_calls_found'],
                                        'total_calls': processing_results['total_calls_found']
                                    })
                                    
                                    # 处理错误
                                    if processing_results['failed_results']:
                                        handle_image_processing_errors(processing_results['failed_results'])
                                    
                                    # 如果有成功处理的结果，保存到数据库
                                    if processing_results['successful_images'] > 0:
                                        progress_placeholder.markdown("**💾 正在保存识别结果到数据库...**")
                                        
                                        # 准备数据库更新数据
                                        db_update_data = prepare_database_update_data(
                                            processing_results, 
                                            st.session_state.salesperson_id
                                        )
                                        
                                        # 保存到数据库
                                        save_success = db_manager.save_image_analysis_data(
                                            st.session_state.salesperson_id,
                                            db_update_data,
                                            st.session_state.upload_choice
                                        )
                                        
                                        if save_success:
                                            progress_placeholder.markdown("**✅ 图片识别结果已成功保存到数据库！**")
                                            
                                            # 保存处理结果到session state用于显示
                                            st.session_state.image_analysis_results = processing_results
                                            st.session_state.analysis_completed = True
                                            
                                            # 显示成功信息
                                            st.success(f"""
                                            ✅ **图片识别完成！**
                                            
                                            📊 **识别结果统计：**
                                            - 处理图片：{processing_results['successful_images']} 张
                                            - 发现通话：{processing_results['total_calls_found']} 个
                                            - 有效通话：{processing_results['effective_calls_found']} 个
                                            """)
                                        else:
                                            st.error("识别结果保存到数据库时出现问题")
                                    else:
                                        st.warning("没有成功识别出任何通话信息，请检查图片质量或内容")
                                    
                                except Exception as e:
                                    st.error(f"图片识别过程中出现错误：{str(e)}")
                                    import traceback
                                    with st.expander("详细错误信息"):
                                        st.code(traceback.format_exc())
                    else:
                        st.warning("⚠️ 没有图片可以处理，请重新选择")

    except Exception as e:
        st.error(f"检查数据库记录时出错：{str(e)}")

if st.session_state.analysis_results:
    # 显示整体转换状态
    conversion_summary = {"total": 0, "converted": 0, "failed": 0, "no_conversion": 0}
    converted_files_info = []
    
    for res in st.session_state.analysis_results:
        if res["status"] == "success":
            conversion_summary["total"] += 1
            if "conversion_info" in res:
                if res["conversion_info"].get("conversion_success", False):
                    conversion_summary["converted"] += 1
                    converted_files_info.append({
                        "filename": os.path.basename(res["file_path"]),
                        "original_size": res["conversion_info"]["original_size_bytes"],
                        "converted_size": res["conversion_info"]["converted_size_bytes"],
                        "duration": res["conversion_info"]["converted_duration_seconds"]
                    })
                else:
                    conversion_summary["failed"] += 1
            else:
                conversion_summary["no_conversion"] += 1
    
    # 显示转换摘要
    if conversion_summary["converted"] > 0 or conversion_summary["failed"] > 0:
        st.markdown("### 🔄 文件转换状态")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("总文件数", conversion_summary["total"])
        with col2:
            st.metric("转换成功", conversion_summary["converted"], delta=f"+{conversion_summary['converted']}" if conversion_summary["converted"] > 0 else None)
        with col3:
            st.metric("转换失败", conversion_summary["failed"], delta=f"-{conversion_summary['failed']}" if conversion_summary["failed"] > 0 else None)
        with col4:
            st.metric("无需转换", conversion_summary["no_conversion"])
        
        # 显示转换成功的文件详情
        if converted_files_info:
            with st.expander(f"📋 转换成功的文件详情 ({len(converted_files_info)} 个)", expanded=False):
                for i, file_info in enumerate(converted_files_info, 1):
                    size_change = file_info["converted_size"] / file_info["original_size"] if file_info["original_size"] > 0 else 1
                    change_text = f"{size_change:.2f}x" if size_change != 1 else "无变化"
                    change_color = "🔻" if size_change < 1 else "🔺" if size_change > 1 else "➖"
                    
                    st.markdown(f"**{i}. {file_info['filename']}**")
                    st.markdown(f"   - 原始大小: {file_info['original_size']:,} 字节")
                    st.markdown(f"   - 转换后大小: {file_info['converted_size']:,} 字节 {change_color} {change_text}")
                    st.markdown(f"   - 音频时长: {file_info['duration']:.2f} 秒")
        
        if conversion_summary["converted"] > 0:
            st.info("💡 转换后的文件已保留供验证，将在程序结束时自动清理")
        
        st.markdown("---")
    
    tab1, tab2, tab3 = st.tabs(["📝 所有对话记录", "📊 所有分析结果", "📈 汇总分析"])

    with tab1:
        for idx, res in enumerate(st.session_state.analysis_results, 1):
            if res["status"] == "success":
                analysis_result = res["analysis_result"]
                if analysis_result.get("status") == "success":
                    st.markdown(f"### 📝 对话记录 {idx}")
                    
                    # 显示转换文件信息（如果有）
                    if "conversion_info" in res:
                        conversion_info = res["conversion_info"]
                        if conversion_info.get("conversion_success", False):
                            with st.expander(f"🔄 文件转换信息 - {os.path.basename(res['file_path'])}", expanded=False):
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.markdown("**原始文件：**")
                                    st.markdown(f"- 文件路径: `{conversion_info['original_file_path']}`")
                                    st.markdown(f"- 文件大小: {conversion_info['original_size_bytes']:,} 字节")
                                    
                                with col2:
                                    st.markdown("**转换后文件：**")
                                    st.markdown(f"- 文件路径: `{conversion_info['converted_file_path']}`")
                                    st.markdown(f"- 文件大小: {conversion_info['converted_size_bytes']:,} 字节")
                                    st.markdown(f"- 音频时长: {conversion_info['converted_duration_seconds']:.2f} 秒")
                                    st.markdown(f"- 音频格式: {conversion_info['converted_format']}")
                                    st.markdown(f"- 采样率: {conversion_info['converted_sample_rate']:,} Hz")
                                    st.markdown(f"- 声道数: {conversion_info['converted_channels']}")
                                
                                # 显示转换效果
                                compression_ratio = conversion_info['original_size_bytes'] / conversion_info['converted_size_bytes'] if conversion_info['converted_size_bytes'] > 0 else 0
                                if compression_ratio > 1:
                                    st.success(f"✅ 转换成功！文件大小变化: {compression_ratio:.2f}x")
                                else:
                                    st.info(f"ℹ️ 转换成功！文件大小变化: {1/compression_ratio:.2f}x (增大)")
                        else:
                            st.error(f"❌ 文件转换失败: {conversion_info.get('conversion_error', '未知错误')}")
                    
                    if analysis_result["roles"].get("confidence", "low") != "high":
                        st.warning("⚠️ 该对话的角色识别可信度不高，请核实。")
                    st.markdown(f"**角色说明：**")
                    st.markdown(f"- 说话者1 ({analysis_result['roles']['spk1']})")
                    st.markdown(f"- 说话者2 ({analysis_result['roles']['spk2']})")
                    st.markdown("**详细对话：**")
                    st.markdown(analysis_result["formatted_text"])
                    st.markdown("---")

    with tab2:
        for idx, res in enumerate(st.session_state.analysis_results, 1):
            if res["status"] == "success":
                analysis_result = res.get("analysis_result", {})
                if analysis_result.get("status") == "success":
                    file_name = os.path.basename(res["file_path"])
                    file_name = re.sub(r'^temp_', '', file_name)
                    file_name = os.path.splitext(file_name)[0]
                    with st.expander(f"📊 {file_name} 通话分析"):
                        st.markdown(analysis_result["analysis"])
                        st.markdown("---")
                else:
                    st.error(f"文件 {idx} 分析失败: {analysis_result.get('message', '未知错误')}")

    with tab3:
        st.markdown("### 📈 汇总分析报告")
        st.markdown(st.session_state.summary_analysis)

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📥 下载完整分析报告",
            data=st.session_state.combined_report,
            file_name="complete_analysis_report.md",
            mime="text/plain"
        )

    with col2:
        excel_filename = "电话开拓分析表_未知联系人.xlsx"  # 默认文件名
        
        def generate_excel_report():
            try:
                workbook = openpyxl.load_workbook(EXCEL_CONFIG["template_file"])
                worksheet = workbook.active
                file_names = []
                contact_persons = []
                analysis_data = []
                for res in st.session_state.analysis_results:
                    if res["status"] == "success" and res["analysis_result"].get("status") == "success":
                        file_name = os.path.basename(res["file_path"])
                        file_name = re.sub(r'^temp_', '', file_name)
                        file_name = os.path.splitext(file_name)[0]
                        
                        # 使用智能文件名解析
                        company_name, contact_person, phone_number = parse_filename_intelligently(file_name)
                        
                        file_names.append(company_name)
                        contact_persons.append(contact_person)
                        
                        # 使用新的精确提取函数
                        analysis_text = res["analysis_result"]["analysis"]
                        extracted_data = extract_all_conversation_data(analysis_text)
                        
                        analysis_data.append({
                            "score": extracted_data["score"], 
                            "suggestion": extracted_data["suggestion"], 
                            "phone_number": phone_number,
                            "contact_person": contact_person
                        })
                
                # 查找表格中的列
                column_indices = {}
                for col in range(1, worksheet.max_column + 1):
                    header = worksheet.cell(1, col).value
                    if header:
                        column_indices[header] = col
                
                # 填写数据到表格中
                for i, (name, data) in enumerate(zip(file_names, analysis_data)):
                    row = i + 2
                    if row <= worksheet.max_row:
                        if "客户名称" in column_indices:
                            worksheet.cell(row, column_indices["客户名称"]).value = name
                        if "联系人" in column_indices:
                            worksheet.cell(row, column_indices["联系人"]).value = data["contact_person"]
                        if "联系电话" in column_indices and data["phone_number"]:
                            worksheet.cell(row, column_indices["联系电话"]).value = data["phone_number"]
                        if "评分" in column_indices and data["score"]:
                            try:
                                worksheet.cell(row, column_indices["评分"]).value = int(data["score"])
                            except ValueError:
                                worksheet.cell(row, column_indices["评分"]).value = data["score"]
                        if "通话优化建议" in column_indices and data["suggestion"]:
                            worksheet.cell(row, column_indices["通话优化建议"]).value = data["suggestion"]
                
                # 填写该日电话数
                total_calls = len([res for res in st.session_state.analysis_results if res["status"] == "success"])
                # 寻找"该日电话数"单元格
                for row in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row, 1).value
                    if cell_value and "该日电话数" in str(cell_value):
                        # 假设CDEF合并单元格在第3列开始
                        worksheet.cell(row, 3).value = total_calls
                        break
                
                # 处理总结部分
                if st.session_state.summary_analysis:
                    # 使用新的精确提取函数
                    summary_data = extract_all_summary_data(st.session_state.summary_analysis)
                    avg_score = summary_data["average_score"]
                    improvement_measures = summary_data["improvement_measures"]
                    
                    # 格式化改进措施
                    formatted_suggestions = ""
                    if improvement_measures:
                        formatted_suggestions = "改进建议：\n"
                        for measure in improvement_measures:
                            formatted_suggestions += f"- {measure}\n"
                    else:
                        # 如果没有提取到措施，使用原始内容的前几行作为备选
                        formatted_suggestions = "改进建议：\n- 请查看详细分析报告"
                    
                    # 找到总结行
                    summary_row = None
                    for row in range(1, worksheet.max_row + 1):
                        cell_value = worksheet.cell(row, 1).value
                        if cell_value and "总结" in str(cell_value):
                            summary_row = row
                            break
                    
                    if not summary_row:
                        # 如果没找到，默认使用第33行
                        summary_row = EXCEL_CONFIG["summary_row"]
                    
                    if formatted_suggestions:
                        worksheet.cell(summary_row, 2).value = formatted_suggestions
                        # 设置改进建议单元格对齐方式：顶部对齐 + 自动换行
                        worksheet.cell(summary_row, 2).alignment = openpyxl.styles.Alignment(
                            wrapText=True, 
                            vertical='top',
                            horizontal='left'
                        )
                    
                    # 查找总评分列
                    total_score_col = None
                    for col in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(summary_row, col).value
                        if cell_value and "总评分" in str(cell_value):
                            total_score_col = col
                            break
                    
                    if total_score_col and avg_score:
                        worksheet.cell(summary_row, total_score_col).value = f"总评分：\n{avg_score}"
                        # 设置单元格对齐方式：顶部对齐 + 自动换行
                        worksheet.cell(summary_row, total_score_col).alignment = openpyxl.styles.Alignment(
                            wrapText=True, 
                            vertical='top',
                            horizontal='left'
                        )
                
                # 获取第一个文件的联系人名称，如果没有则使用默认值
                first_contact = contact_persons[0] if contact_persons and contact_persons[0] else "未知联系人"
                
                # 获取当前日期
                from datetime import datetime
                today_date = datetime.now().strftime("%Y%m%d")
                
                # 生成文件名
                global excel_filename
                excel_filename = f"电话开拓分析表_{first_contact}_{today_date}.xlsx"
                
                # 保存到内存中
                excel_buffer = BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            except Exception as e:
                logging.error(f"生成Excel报告时出错: {e}")
                st.error(f"生成Excel报告时出错: {e}")
                return None

        excel_data = generate_excel_report()
        if excel_data:
            st.download_button(
                label="📊 下载电话开拓分析表",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# 显示图片识别结果（新增）
elif hasattr(st.session_state, 'image_analysis_results') and st.session_state.image_analysis_results:
    st.markdown("### 📸 图片识别结果")
    
    results = st.session_state.image_analysis_results
    
    # 显示统计摘要
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("处理图片", results['total_images'])
    with col2:
        st.metric("识别成功", results['successful_images'])
    with col3:
        st.metric("发现通话", results['total_calls_found'])
    with col4:
        st.metric("有效通话", results['effective_calls_found'])
    
    st.markdown("---")
    
    # 显示详细的通话记录
    if results['all_calls']:
        st.markdown("### 📞 识别到的通话记录")
        
        for idx, call in enumerate(results['all_calls'], 1):
            with st.expander(f"📞 通话记录 {idx} - {call.get('contact_info', '未知联系人')}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**基本信息：**")
                    st.markdown(f"- 联系人：{call.get('contact_info', '未知')}")
                    st.markdown(f"- 公司：{call.get('company_name', '未知')}")
                    st.markdown(f"- 通话时间：{call.get('call_time', '未知')}")
                
                with col2:
                    st.markdown("**通话统计：**")
                    duration_text = call.get('duration_text', '未知')
                    duration_seconds = call.get('duration_seconds', 0)
                    is_effective = call.get('is_effective', False)
                    
                    st.markdown(f"- 通话时长：{duration_text} ({duration_seconds}秒)")
                    
                    if is_effective:
                        st.success("✅ 有效通话")
                    else:
                        st.warning("⚠️ 无效通话")
                
                # 显示附加信息
                if call.get('additional_info'):
                    st.markdown("**附加信息：**")
                    st.markdown(call['additional_info'])
    
    # 显示处理错误（如果有）
    if results['failed_results']:
        st.markdown("### ⚠️ 处理失败的图片")
        for error in results['failed_results']:
            st.error(f"**{error['filename']}**: {error['error']}")
